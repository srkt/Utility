"""
Debug script to inspect Excel headers and column mapping.
Run this to see exactly what headers are being read
and whether they match the column mapping.

Usage: python debug_headers.py path/to/your/file.xlsx
"""
import sys
import re
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

import openpyxl
import yaml
from etl.extractor import clean_header

# Load config
config_path = Path(__file__).parent / "config" / "pipeline_config.yaml"
with open(config_path) as f:
    CONFIG = yaml.safe_load(f)

EXCEL_CONFIG = CONFIG["excel"]


def debug_file(file_path: str, sheet_key: str = "savings_results"):
    """
    Read Excel headers and show:
    1. Raw header values from all 3 rows
    2. Cleaned header values
    3. Whether they match column mapping
    """
    file_path = Path(file_path)
    if not file_path.exists():
        print(f"ERROR: File not found: {file_path}")
        return

    sheet_config  = EXCEL_CONFIG["sheets"][sheet_key]
    sheet_name    = sheet_config["sheet_name"]
    header_row    = sheet_config["header_row"]
    row1_idx      = sheet_config["row1_header_row"]
    row2_idx      = sheet_config["row2_header_row"]

    print(f"\n{'='*70}")
    print(f"FILE  : {file_path.name}")
    print(f"SHEET : {sheet_name}")
    print(f"{'='*70}")

    # Open workbook
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found")
        print(f"Available sheets: {wb.sheetnames}")
        return

    ws = wb[sheet_name]
    all_rows = list(ws.values)

    row1_raw = all_rows[row1_idx - 1]
    row2_raw = all_rows[row2_idx - 1]
    row3_raw = all_rows[header_row - 1]

    # Propagate merged cells
    def propagate(row):
        result = []
        last = ""
        for val in row:
            if val is not None and str(val).strip() != "":
                last = str(val).strip()
            result.append(last)
        return result

    row1 = propagate(row1_raw)
    row2 = propagate(row2_raw)

    print(f"\n{'─'*70}")
    print(f"{'COL':>4} | {'ROW1 (raw)':20} | {'ROW2 (raw)':25} | {'ROW3 (raw)':30}")
    print(f"{'─'*70}")

    for i, r3 in enumerate(row3_raw):
        if r3 is None:
            continue
        r1 = row1[i] if i < len(row1) else ""
        r2 = row2[i] if i < len(row2) else ""
        print(f"{i:>4} | {str(r1)[:20]:20} | {str(r2)[:25]:25} | {str(r3)[:30]:30}")

    print(f"\n{'─'*70}")
    print("CLEANED HEADERS + MAPPING MATCH")
    print(f"{'─'*70}")
    print(f"{'COL':>4} | {'CLEAN R1':20} | {'CLEAN R2':25} | {'CLEAN R3':35} | {'ORACLE COL':35} | {'MATCH'}")
    print(f"{'─'*70}")

    # Load column mapping
    try:
        from utils.db import db
        mappings = db.load_column_mapping(sheet_name)
        use_db = True
        print("(Using Oracle column mapping)")
    except Exception:
        # Fall back to YAML
        use_db = False
        config_map_path = Path(__file__).parent / "config" / "column_mapping.yaml"
        with open(config_map_path) as f:
            yaml_map = yaml.safe_load(f)
        mappings = yaml_map.get(
            "savings_results" if sheet_key == "savings_results"
            else "savings_forecast", []
        )
        # Normalize yaml keys to uppercase
        mappings = [
            {
                "EXCEL_ROW1_HEADER": m.get("excel_row1", ""),
                "EXCEL_ROW2_HEADER": m.get("excel_row2", ""),
                "EXCEL_ROW3_HEADER": m.get("excel_row3", ""),
                "STG_COLUMN_NAME":   m.get("stg_column", "")
            }
            for m in mappings
        ]
        print("(Using YAML column mapping - Oracle not connected)")

    # Build lookup
    lookup = {}
    for m in mappings:
        key = (
            clean_header(m.get("EXCEL_ROW1_HEADER", "")),
            clean_header(m.get("EXCEL_ROW2_HEADER", "")),
            clean_header(m.get("EXCEL_ROW3_HEADER", ""))
        )
        lookup[key] = m["STG_COLUMN_NAME"]

    # Check each column
    mapped_count   = 0
    unmapped_count = 0
    unmapped_cols  = []

    for i, r3 in enumerate(row3_raw):
        if r3 is None or str(r3).strip() == "":
            continue

        r1_val = row1[i] if i < len(row1) else ""
        r2_val = row2[i] if i < len(row2) else ""

        cr1 = clean_header(r1_val)
        cr2 = clean_header(r2_val)
        cr3 = clean_header(r3)

        if not cr3:
            continue

        # Try full key
        oracle_col = lookup.get((cr1, cr2, cr3))
        match_type = "FULL"

        # Try without row1
        if not oracle_col:
            oracle_col = lookup.get(("", cr2, cr3))
            match_type = "NO_R1"

        # Try without row1 and row2
        if not oracle_col:
            oracle_col = lookup.get(("", "", cr3))
            match_type = "DIM"

        if oracle_col:
            mapped_count += 1
            status = f"✓ {match_type}"
        else:
            unmapped_count += 1
            status = "✗ NO MATCH"
            unmapped_cols.append({
                "col": i,
                "r1": r1_val, "r2": r2_val, "r3": r3,
                "cr1": cr1, "cr2": cr2, "cr3": cr3
            })
            oracle_col = "---"

        print(
            f"{i:>4} | {cr1[:20]:20} | {cr2[:25]:25} | "
            f"{cr3[:35]:35} | {str(oracle_col)[:35]:35} | {status}"
        )

    print(f"\n{'='*70}")
    print(f"SUMMARY: {mapped_count} mapped, {unmapped_count} unmapped")
    print(f"{'='*70}")

    if unmapped_cols:
        print(f"\n{'─'*70}")
        print("UNMAPPED COLUMNS - need to fix in CES_SAVINGS_COLUMN_MAP:")
        print(f"{'─'*70}")
        for u in unmapped_cols:
            print(f"\nCol {u['col']}:")
            print(f"  Raw  : R1='{u['r1']}' | R2='{u['r2']}' | R3='{u['r3']}'")
            print(f"  Clean: R1='{u['cr1']}' | R2='{u['cr2']}' | R3='{u['cr3']}'")
            print(f"  Fix  : Add to mapping with stg_column name")
    else:
        print("\n✓ All columns mapped successfully!")

    wb.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python debug_headers.py <path_to_excel_file> [sheet_key]")
        print("  sheet_key: savings_results (default) or savings_forecast")
        print("\nExample:")
        print("  python debug_headers.py C:/ces_savings/landing/ICF_2026_04.xlsx")
        print("  python debug_headers.py C:/ces_savings/landing/ICF_2026_04.xlsx savings_forecast")
        sys.exit(1)

    file_arg  = sys.argv[1]
    sheet_arg = sys.argv[2] if len(sys.argv) > 2 else "savings_results"

    debug_file(file_arg, sheet_arg)

    # Also run forecast sheet if not specified
    if len(sys.argv) < 3:
        print("\n")
        debug_file(file_arg, "savings_forecast")
